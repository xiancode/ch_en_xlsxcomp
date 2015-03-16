#!/usr/bin/python
#encoding=utf-8

import os
import os.path
import string
import logging
from openpyxl import load_workbook
import openpyxl.cell as ce

LOG_FILENAME = 'log.txt'
FORMAT  = '%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s'
DATEFMT = '%a,%d %b %Y %H:%M:%S' 
logging.basicConfig(format = FORMAT,datefmt = DATEFMT,filename = LOG_FILENAME,filemode = 'w',level=logging.DEBUG)

    
def value_change(value):
    """
    对单元格的值进行转化
    """
    result = ""
    if type(value) == unicode:
        result = string.strip(value)
    else:
        result = str(value)
        result = string.strip(result)
    return result

def del_none_row(l):
    """
    删除二维数组中空白的行
    """
    empty_row_num = []
    for row_num in range(len(l)):
        count = 0
        for value in l[row_num]:
            if value == None or value_change(value) == "":
                count += 1
        if count < len(l[row_num]):
            empty_row_num.append(row_num)
    result = []
    for row_num in empty_row_num:
        result.append(l[row_num])
    return result

def get_dims(l):
    """
    
    """
    row_size = len(l)
    col_size = len(l[0])
    for row in l:
        if col_size == len(row):
            pass
        else:
            return [-1,-1]
    return [row_size,col_size]
        
            
class XlsxTableHeader:
    """
    
    """
    def __init__(self,filename):
        try:
            self.filename = filename
            self.wb = load_workbook(filename)
            #self.sname = filename[-5]+"_tmp.xlsx"
            self.ws = self.wb.active
            #最大行
            self.max_col  = self.ws.get_highest_column()
            #最大列
            self.max_row  = self.ws.get_highest_row()
            #
            self.get_top_part_row()
            self.get_theader_row()
            self.get_btm_part_row()
            self.get_last_col()
            self.get_theader_range()
            self.col_row_range()
            #默认从第一列开始
            self.theader_start_col_num = 1
            self.start = True
            msg = "表头区域：" + str(self.theader_range)
            #print "表头区域是",self.theader_range
            logging.info(msg)
            
        except:
            print "初始化",filename,"失败"
            msg = "初始化" + filename + "失败"
            logging.error(msg)
            self.start = False
    def get_theader_content(self):
        """
        归并表头单元格内容，获取最终归并的单元格的值
        """
        #获取表头中合并的区域
        self.get_merged_range()
        marea_list = self.get_theader_merged_ranges()
        self.merged_cells_check(self.ws,marea_list)
        #获取未合并的区域
        nmcells = self.get_not_merged_cells()
        canmcells = self.get_can_merged_range(nmcells)
        self.merge_range(self.ws,canmcells)
        cont = self.get_content()
        #self.wb.save(self.sname)
        return cont
        
            
    def get_top_part_row(self):
        """
        获取首列单元格顶部有边框线的行号
        """
        self.top_partition_row = []
        for i in range(1,self.max_row+1):
            cur_cell = self.ws['A'+str(i)]
            if cur_cell.style.border.top.style:
                self.top_partition_row.append(i)
        return self.top_partition_row
    
    def get_theader_row(self):
        """
        根据顶部边框线来确定表头的起始、结束行号
        """
        self.theader_start_row = self.top_partition_row[0]
        self.theader_end_row   = self.top_partition_row[1]-1
    
    def get_btm_part_row(self):
        """
        获取首列单元格底部有边框线的行号
        """
        self.btm_partition_row = []
        for i in range(1,self.max_row+1):
            cur_cell = self.ws['A'+str(i)]
            if cur_cell.style.border.bottom.style:
                self.btm_partition_row.append(i)
        return self.btm_partition_row
    
    def get_last_col(self):
        """
        根据线条来确定最后一列的列号及列字母
        """
        #从表头首行、末列开始遍历，寻找单元格顶部分割线与首列一样的
        self.theader_end_col = ""
        #theader_top_line = self.ws['A'+str(self.theader_start_row)].style.border.top.style
        for i in range(1,self.max_col+1)[::-1]:
            cur_cell = self.ws[ce.get_column_letter(i)+str(self.theader_end_row+1)]
            #if cur_cell.style.border.top.style == theader_top_line:
            if cur_cell.value != None:
                self.theader_end_col = ce.get_column_letter(i)
                self.theader_end_col_num = i
                break
        return self.theader_end_col
    
    def get_theader_range(self):
        """
        根据表头行号和列号，确定表头区域
        """
        self.theader_range = ""
        self.theader_range = 'A'+str(self.theader_start_row)+":"+self.theader_end_col+str(self.theader_end_row)
        return self.theader_range
    
    def col_row_range(self):
        """
        获取行号和列号的范围 A4:Q7 -> [4,7] [A,Q]
        """
        cell_start,cell_end = self.theader_range.split(":")
        min_col,min_row = ce.column_index_from_string(self.ws[cell_start].column),self.ws[cell_start].row
        max_col,max_row = ce.column_index_from_string(self.ws[cell_end].column),self.ws[cell_end].row
        self.rownum_range = [min_row,max_row]
        self.colnum_range = [ce.get_column_letter(min_col),ce.get_column_letter(max_col)]
        #表头区域大小
        self.theader_area_size = [max_row-min_row+1,max_col-min_col+1]
        #pos = self.theader_range.find(":")
        #if len(self.theader_range[pos+1:]) == 2:
        #    self.colnum_range = [self.theader_range[0],self.theader_range[3]]
        #    self.rownum_range = [self.theader_range[1],self.theader_range[4]]
        #elif len(self.theader_range[pos+1:]) == 3:
        #    self.colnum_range = [self.theader_range[0],self.theader_range[3:5]]
        #    self.rownum_range = [self.theader_range[1],self.theader_range[5]]
    
    def get_merged_range(self):
        """
        获取表格中所有合并单元格的区域
        """
        self.merged_ranges = []
        all_merged_ranges = self.ws.merged_cell_ranges
        for tmp in all_merged_ranges:
            if tmp.find('XFD') == -1:
                self.merged_ranges.append(tmp)
        return self.merged_ranges
    
    def get_theader_merged_ranges(self):
        """
        获取表头中的合并区域
        """
        self.theader_merged_ranges = []
        for cell_range in self.merged_ranges:
            cell_start,cell_end = cell_range.split(":") 
            #判断区域的起始 和 结尾单元格是否属于表头
            #print ce.column_index_from_string(self.ws[cell_start].column)
            if self.theader_start_col_num <= ce.column_index_from_string(self.ws[cell_start].column) <= self.theader_end_col_num \
               and self.rownum_range[0] <= self.ws[cell_start].row <= self.rownum_range[1] \
               and self.theader_start_col_num <= ce.column_index_from_string(self.ws[cell_end].column) <= self.theader_end_col_num \
               and self.rownum_range[0] <= self.ws[cell_end].row <= self.rownum_range[1]:
                self.theader_merged_ranges.append(cell_range)

        self.theader_merged_ranges.sort()
        return self.theader_merged_ranges
    
    def get_all_cells(self,cell_ranges):
        """
        根据给定的区域，获取该区域中的所有单元格 A4:A5
        """
        result_cells = []
        min_col = 0
        min_row = 0
        max_col = 0
        max_row = 0
        cell_start,cell_end = cell_ranges.split(":")
        min_col,min_row = ce.column_index_from_string(self.ws[cell_start].column),self.ws[cell_start].row
        max_col,max_row = ce.column_index_from_string(self.ws[cell_end].column),self.ws[cell_end].row
        for c_t in range(min_col,max_col+1):
            for r_t in range(min_row,max_row+1):
                result_cells.append(ce.get_column_letter(c_t)+str(r_t))
        return result_cells
    
    def get_not_merged_cells(self):
        """
        获取表头中没有合并的单元格
        """
        theader_all_cells = self.get_all_cells(self.theader_range)
        not_merged_cells_set = set(theader_all_cells)
        for cell_range in self.theader_merged_ranges:
            not_merged_cells_set = not_merged_cells_set - set(self.get_all_cells(cell_range))
        self.not_merged_cells = list(not_merged_cells_set)
        return self.not_merged_cells
    
    def get_max_range(self,cells_range):
        """
        获取包含合并单元格的最大区域 ['AA5:AE5']
        """
        cells = []
        min_col = 0
        min_row = 0
        max_col = 0
        max_row = 0
        for cell_range in cells_range:
                cells += cell_range.split(":")
                
        min_col = max_col = ce.column_index_from_string(self.ws[cells[0]].column)
        min_row = max_row = self.ws[cells[0]].row
        for cur_cell in cells:
            if ce.column_index_from_string(self.ws[cur_cell].column) <= min_col:
                min_col = ce.column_index_from_string(self.ws[cur_cell].column)
            if ce.column_index_from_string(self.ws[cur_cell].column) >= max_col:
                max_col = ce.column_index_from_string(self.ws[cur_cell].column)
            if self.ws[cur_cell].row <= min_row:
                min_row = self.ws[cur_cell].row
            if self.ws[cur_cell].row >= max_row:
                max_row = self.ws[cur_cell].row  
        return ce.get_column_letter(min_col) + str(min_row) + ":" + ce.get_column_letter(max_col) + str(max_row)
    
    def get_can_merged_range(self,not_merged_cells_list):
        """
        获取还可以合并的未合并单元格
        """
        result = [] 
        tmp_set = set(not_merged_cells_list)
        while tmp_set:
            cur_cell = tmp_set.pop()
            #print cur_cell
            samecol_set = set()
            #samecol_set.add(cell)
            for cell_ in tmp_set:
                if self.ws[cell_].column == self.ws[cur_cell].column:
                    samecol_set.add(cell_)
                #print cell,
            tmp_set = tmp_set - samecol_set
            samecol_set.add(cur_cell)
            samecol_list = list(samecol_set)
            #保证行的递增顺序
            samecol_list.sort()
            result.append(samecol_list)
        return result
    
    def merged_cells_check(self,ws,merged_areas_list):
        """
        合并单元格的检测,是否值在第一个单元格
        """
        for merged_area in merged_areas_list:
            merged_cells = self.get_all_cells(merged_area)
            scell_num = merged_cells[0]
            #不用把单元格值转化为str 然后在进行string.strip() ,可以直接对单元格值进行string.strip()
            if ws[scell_num].value != None and value_change(ws[scell_num].value) != "":
                pass
            else:
                for idx in range(1,len(merged_cells)):
                    cell_num = merged_cells[idx]
                    #if ws[cell_num].value != None and string.strip(ws[cell_num].value) != "":
                    if ws[cell_num].value != None and value_change(ws[cell_num].value) != "":
                        ws[scell_num].value = ws[cell_num].value
                        ws[cell_num].value = None

  
    def merge_range(self,ws,can_merged_range):
        """
        
        """
        for item in can_merged_range:
            if len(item) == 1:
                continue;
            #print item[0]+":"+item[-1]
            merge_range = self.get_max_range([item[0]+":"+item[-1]])
            #判断列是否相同
            if merge_range[0] == merge_range[3]:
                all_cells = self.get_all_cells(merge_range)
                #print "all_cells",all_cells
                #
                for cell in all_cells:
                    if ws[cell].value == None:
                        ws[cell].value = ""
                if len(all_cells) >= 2:
                    #只有两行，判断第一行的底部 和 第二行的顶部是否有边框，没有边框则合并
                    if len(all_cells) == 2:
                        if ws[all_cells[0]].style.border.top.style and ws[all_cells[-1]].style.border.bottom.style \
                            and ws[all_cells[0]].style.border.bottom.style == None  \
                            and ws[all_cells[-1]].style.border.top.style == None:
                            #print "合并前",ws[all_cells[0]].value,ws[all_cells[-1]].value
                            ws[all_cells[0]].value = value_change(ws[all_cells[0]].value)  + " " + value_change(ws[all_cells[-1]].value)
                            ws[all_cells[-1]].value = None
                            #print all_cells,"单元格值合并成功"
                            #print "合并后",ws[all_cells[0]].value,ws[all_cells[-1]].value
                        else:
                            #print all_cells,"单元格值不能合并"
                            msg = self.filename + str(all_cells) + "单元格值不能合并"
                            logging.info(msg)
                    #超过两行，判断中间行是否包含边框
                    else:
                        if ws[all_cells[0]].style.border.top.style and ws[all_cells[-1]].style.border.bottom.style:
                            merge_flag = 1
                            for idx in range(1,len(all_cells)-1):
                                #如果有一个单元格存在上下边框线，则不能合并
                                if ws[all_cells[idx]].style.border.top.style or ws[all_cells[idx]].style.border.bottom.style :
                                    merge_flag = 0;
                                    #print all_cells,"单元格值不能合并，有边框"
                                    msg = self.filename + str(all_cells) + "单元格值不能合并，有边框"
                                    logging.info(msg)
                                    break
                            if merge_flag ==1 :
                                for idx in range(1,len(all_cells))[::-1]:
                                    ws[all_cells[idx-1]].value = value_change(ws[all_cells[idx-1]].value) + " " + value_change(ws[all_cells[idx]].value) 
                                    ws[all_cells[idx]].value  = ""
                                #print all_cells,"单元格值合并成功！"   
                else:
                    #print "行数少于两行 error"
                    msg = self.filename + "行数少于两行"
                    logging.info(msg)
    
    def get_content(self):
        """
        获取单元格内容，按照二维数组存储每个单元格的内容
        """
        self.content = []
        for i in range(self.rownum_range[0],self.rownum_range[1]+1):
            tmp_list = []
            for j in range(self.theader_start_col_num,self.theader_end_col_num+1):
                tmp_list.append(self.ws[ce.get_column_letter(j)+str(i)].value)
            self.content.append(tmp_list)
        return self.content
        
#主函数
def ch_en_xlsxcomp():
    """
    根据给定的中英文对应表格，完成中英文表头的对应输出
    """
    fout = open("TableHeaderContent.txt",'a')
    #flog = open("log.txt","w")
    conf_file = open("conf.data","r")
    conf_lines = conf_file.readlines()
    ch_dir = ""
    en_dir = ""
    if len(conf_lines) == 2:
        ch_dir = conf_lines[0].strip()
        en_dir = conf_lines[1].strip()
    else:
        print "conf.data error!"
    #获取文件名    
    filename_list = []
    for parent,dirnames,filenames in os.walk(ch_dir):
        for filename in filenames:
            filename_list.append(filename)
    
    #
    #for i in range(1):
    for filename in filename_list:
        #filename = "A0111c.xlsx"
        ch_xlsx_name = ch_dir + filename
        en_xlsx_name = en_dir + filename
        print "当前处理文件号：",filename
        msg = "当前处理文件号:" + filename
        logging.info(msg)
        #flog.write(filename+"\n")
        ch_xl = XlsxTableHeader(ch_xlsx_name)
        en_xl = XlsxTableHeader(en_xlsx_name)
        
        if ch_xl.start and en_xl.start:
            ch_content = ch_xl.get_theader_content()
            en_content = en_xl.get_theader_content()
            ch_content = del_none_row(ch_content)
            en_content = del_none_row(en_content)
            if get_dims(ch_content) != [-1,-1]  and get_dims(en_content) != [-1,-1] and get_dims(ch_content) == get_dims(en_content):
                fout.write(filename+"----------\n")
                for idx_row in range(len(ch_content)):
                    for idx_col in range(len(ch_content[idx_row])):
                        if (ch_content[idx_row][idx_col] != None and ch_content[idx_row][idx_col] != "") \
                            or (en_content[idx_row][idx_col] != None and en_content[idx_row][idx_col] != ""):
                            #try:
                            ch_ = ch_content[idx_row][idx_col]
                            en_ = en_content[idx_row][idx_col]
                            if type(ch_) == unicode:
                                ch_ = ch_.replace("\n"," ")
                            if type(en_) == unicode:
                                en_ = en_.replace("\n"," ")
                            if ch_ == None:
                                ch_ = u"None"
                            if en_ == None:
                                en_ = u"None"
                                #print ch_,"||",en_
                                #fout.write(ch_.encode('UTF-8'))
                            fout.write(value_change(ch_).encode('UTF-8'))
                            fout.write("||")
                            #fout.write(en_.encode('UTF-8'))
                            fout.write(value_change(en_).encode('UTF-8'))
                            fout.write("\n")
                            #except:
                                #fout.write(ch_.encode('UTF-8'))
                                #fout.write("value error \n")
                        else:
                            pass    
            else:
                #fout.write("表头区域大小不一致\n")
                msg = filename + "表头区域不一致"
                logging.info(msg)                    
        else:
            fout.write("文件初始化没有完成\n")
            msg = filename + "文件初始化没有完成"
            logging.info(msg)
            print "文件",filename,"有错误，请检查"
    fout.close()
    #flog.close()
    print "Compare End!"


if __name__ == "__main__":
    ch_en_xlsxcomp()
    #ch_xl = XlsxTableHeader("./ch_xlsx/B0903a.xlsx")
    #if ch_xl.start:
    #    ch_xl.get_theader_content()
    #    for line in ch_xl.content:
    #        for item in line:
    #            print item,"|",
    #        print "-->"
        #print "---"
    print "End!"
    
    
        
    
        
    
    